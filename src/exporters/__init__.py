"""Turning a report into a file the manager can open (TZ 5.9).

Format only. Nothing here knows what a shift or a checklist *is* — it is handed the frozen
snapshot from `src/services/reports.py` and lays it out; the headings come from
`src/bot/texts/reports.py`, because they are wording and the customer edits wording without
a developer (TZ 8.2).

**Why `.xlsx` in the chat and not a link to a spreadsheet in the cloud.** TZ 5.9 asks for a
file, it needs no account anywhere, it carries no personal data outside the perimeter — and
a file Telegram delivers opens in Google Sheets from the chat in two taps, which is what
was actually wanted. The cloud path was costed separately and is a question for the
customer, not a default.

**Building is synchronous and belongs in a thread.** openpyxl is CPU-bound and the bot is
one process: a month assembled on the event loop freezes every employee of the venue for
the duration. Callers use `asyncio.to_thread`; :func:`shift_report_workbook` returns bytes
so that nothing here touches the filesystem — the delivery is `BufferedInputFile`, and the
guard on seeded data forbids writing a workbook into the delivery anyway.
"""

from __future__ import annotations

import datetime as dt
from collections.abc import Iterable, Sequence
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.bot import texts
from src.services.reports import PendingLine, RunLine, ShiftLine, ShiftReport

#: Widths are set from the content because a column of names at the default width shows
#: «####» or a truncated surname, and the first thing a manager does otherwise is drag the
#: borders. Capped: one very long checklist line must not push the sheet off the screen.
MIN_WIDTH = 10
MAX_WIDTH = 60
WIDTH_PADDING = 2


def shift_report_workbook(report: ShiftReport) -> bytes:
    """The report as a workbook, in memory (TZ 5.9).

    Three sheets, and the order is the order of the manager's question: who worked, how the
    checklists ended, and what nobody ticked. The third is the one the report is opened for,
    so it is a sheet of its own rather than a column somewhere.
    """
    book = Workbook()
    _fill(book.active, texts.REPORT_SHIFTS_SHEET, _shift_headings(), _shift_rows(report.shifts))
    _fill(
        book.create_sheet(),
        texts.REPORT_RUNS_SHEET,
        _run_headings(),
        _run_rows(report.runs),
    )
    _fill(
        book.create_sheet(),
        texts.REPORT_PENDING_SHEET,
        _pending_headings(),
        _pending_rows(report.pending),
    )
    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()


def shift_report_filename(report: ShiftReport) -> str:
    """`smena-2026-08-15.xlsx` — sortable, and safe on every filesystem.

    ISO and not the venue-local `15.08`: a folder of these sorts itself by date, which is
    what a manager keeping a month of them needs. No venue name in it — that is the venue's
    own text and would arrive with spaces, quotes and slashes in it.
    """
    return texts.REPORT_FILENAME_TEMPLATE.format(day=report.day.isoformat())


# --------------------------------------------------------------------------------------
# Sheets
# --------------------------------------------------------------------------------------


def _shift_headings() -> tuple[str, ...]:
    return (
        texts.REPORT_COLUMN_EMPLOYEE,
        texts.REPORT_COLUMN_WINDOW,
        texts.REPORT_COLUMN_OPENS,
        texts.REPORT_COLUMN_CLOSES,
    )


def _shift_rows(lines: Sequence[ShiftLine]) -> Iterable[tuple[Any, ...]]:
    for line in lines:
        yield (
            line.full_name,
            texts.REPORT_WINDOW_TEMPLATE.format(
                start=_clock(line.start),
                end=_clock(line.end),
            ),
            _mark(line.is_opener),
            _mark(line.is_closer),
        )


def _run_headings() -> tuple[str, ...]:
    return (
        texts.REPORT_COLUMN_CHECKLIST,
        texts.REPORT_COLUMN_EMPLOYEE,
        texts.REPORT_COLUMN_STATUS,
        texts.REPORT_COLUMN_PROGRESS,
        texts.REPORT_COLUMN_FINISHED,
        texts.REPORT_COLUMN_WHY,
    )


def _run_rows(lines: Sequence[RunLine]) -> Iterable[tuple[Any, ...]]:
    for line in lines:
        yield (
            texts.checklist_title(line.checklist),
            line.full_name,
            texts.run_status_label(line.status),
            texts.REPORT_PROGRESS_TEMPLATE.format(done=line.done, total=line.total),
            _moment(line.completed_at),
            line.skip_comment or "",
        )


def _pending_headings() -> tuple[str, ...]:
    return (
        texts.REPORT_COLUMN_CHECKLIST,
        texts.REPORT_COLUMN_EMPLOYEE,
        texts.REPORT_COLUMN_GROUP,
        texts.REPORT_COLUMN_WORDING,
        texts.REPORT_COLUMN_CRITICAL,
    )


def _pending_rows(lines: Sequence[PendingLine]) -> Iterable[tuple[Any, ...]]:
    for line in lines:
        yield (
            texts.checklist_title(line.checklist),
            line.full_name,
            line.group_name or "",
            line.text,
            _mark(line.is_critical),
        )


# --------------------------------------------------------------------------------------
# Layout
# --------------------------------------------------------------------------------------


def _fill(
    sheet: Worksheet | Any,
    title: str,
    headings: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> None:
    sheet.title = title
    sheet.append(list(headings))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    written = 0
    for row in rows:
        sheet.append(list(row))
        written += 1
    # The heading row is frozen so that scrolling forty unticked lines keeps the columns
    # readable — the sheet is meant to be read, not just archived.
    sheet.freeze_panes = "A2"
    if written:
        sheet.auto_filter.ref = sheet.dimensions
    _fit_columns(sheet, headings)


def _fit_columns(sheet: Worksheet | Any, headings: Sequence[str]) -> None:
    for index in range(1, len(headings) + 1):
        longest = max(
            (len(str(cell.value)) for cell in sheet[get_column_letter(index)] if cell.value),
            default=0,
        )
        width = min(MAX_WIDTH, max(MIN_WIDTH, longest + WIDTH_PADDING))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _clock(value: dt.time) -> str:
    return f"{value.hour:02d}:{value.minute:02d}"


def _moment(value: dt.datetime | None) -> str:
    """A finished-at stamp, or nothing at all for a checklist that never finished."""
    return "" if value is None else value.strftime("%Y-%m-%d %H:%M")


def _mark(flag: bool) -> str:
    return texts.REPORT_YES if flag else ""


__all__ = [
    "MAX_WIDTH",
    "MIN_WIDTH",
    "shift_report_filename",
    "shift_report_workbook",
]
