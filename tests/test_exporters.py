"""The workbook the manager opens (TZ 5.9).

Assertions are made by *reading the file back*, not by trusting the calls that wrote it: a
report that builds without raising and opens to an empty sheet is the failure worth
catching, and only a round trip through openpyxl sees it.
"""

from __future__ import annotations

import datetime as dt
from io import BytesIO

from openpyxl import load_workbook
from src.bot import texts
from src.db.models import ChecklistType, RunStatus
from src.exporters import shift_report_filename, shift_report_workbook
from src.services.reports import PendingLine, RunLine, ShiftLine, ShiftReport

DAY = dt.date(2026, 8, 15)


def make_report(**kwargs: object) -> ShiftReport:
    base: dict[str, object] = {
        "venue": "PIMS",
        "day": DAY,
        "shifts": (
            ShiftLine(
                full_name="Иван Петров",
                start=dt.time(8),
                end=dt.time(14),
                is_opener=True,
                is_closer=False,
            ),
        ),
        "runs": (
            RunLine(
                checklist=ChecklistType.OPENING,
                full_name="Иван Петров",
                status=RunStatus.SKIPPED,
                done=1,
                total=2,
                completed_at=dt.datetime(2026, 8, 15, 8, 20, tzinfo=dt.UTC),
                skip_comment="не привезли лёд",
            ),
        ),
        "pending": (
            PendingLine(
                checklist=ChecklistType.OPENING,
                full_name="Иван Петров",
                text="Проверить лёд",
                group_name="Станция",
                is_critical=True,
            ),
        ),
    }
    base.update(kwargs)
    return ShiftReport(**base)  # type: ignore[arg-type]


def read(report: ShiftReport) -> dict[str, list[tuple[object, ...]]]:
    book = load_workbook(BytesIO(shift_report_workbook(report)))
    return {name: list(book[name].iter_rows(values_only=True)) for name in book.sheetnames}


def test_the_workbook_has_the_three_sheets_of_the_question() -> None:
    """Who worked, how the checklists ended, and what nobody ticked (TZ 5.9)."""
    sheets = read(make_report())

    assert list(sheets) == [
        texts.REPORT_SHIFTS_SHEET,
        texts.REPORT_RUNS_SHEET,
        texts.REPORT_PENDING_SHEET,
    ]


def test_every_sheet_carries_its_headings() -> None:
    sheets = read(make_report())

    assert sheets[texts.REPORT_SHIFTS_SHEET][0][0] == texts.REPORT_COLUMN_EMPLOYEE
    assert sheets[texts.REPORT_RUNS_SHEET][0][0] == texts.REPORT_COLUMN_CHECKLIST
    assert sheets[texts.REPORT_PENDING_SHEET][0][-1] == texts.REPORT_COLUMN_CRITICAL


def test_what_was_not_ticked_reaches_the_file() -> None:
    """The whole point of the report, and the one row that must never be dropped."""
    sheets = read(make_report())

    (_, line) = sheets[texts.REPORT_PENDING_SHEET]
    assert line[2] == "Станция"
    assert line[3] == "Проверить лёд"
    assert line[4] == texts.REPORT_YES, "the critical mark is a word, so the sheet filters on it"


def test_the_progress_and_the_reason_reach_the_file() -> None:
    sheets = read(make_report())

    (_, line) = sheets[texts.REPORT_RUNS_SHEET]
    assert line[2] == texts.run_status_label(RunStatus.SKIPPED)
    assert line[3] == texts.REPORT_PROGRESS_TEMPLATE.format(done=1, total=2)
    assert line[5] == "не привезли лёд"


def test_a_day_with_no_pending_lines_still_has_the_sheet() -> None:
    """A sheet with only headings says «nothing was missed»; a missing sheet says nothing."""
    sheets = read(make_report(pending=()))

    assert len(sheets[texts.REPORT_PENDING_SHEET]) == 1


def test_a_checklist_that_never_finished_leaves_the_stamp_empty() -> None:
    """An unfinished run has no `completed_at`, and a made-up one would be a lie."""
    report = make_report(
        runs=(
            RunLine(
                checklist=ChecklistType.OPENING,
                full_name="Иван Петров",
                status=RunStatus.OVERDUE,
                done=0,
                total=2,
                completed_at=None,
                skip_comment=None,
            ),
        )
    )

    (_, line) = read(report)[texts.REPORT_RUNS_SHEET]

    assert line[4] in (None, "")


def test_the_file_is_named_by_an_iso_date() -> None:
    """A folder of these sorts itself by date, which is what a month of them needs."""
    assert shift_report_filename(make_report()) == "smena-2026-08-15.xlsx"


def test_the_venue_name_is_not_in_the_file_name() -> None:
    """It is the venue's own text and arrives with spaces, quotes and slashes in it."""
    assert "PIMS" not in shift_report_filename(make_report(venue="Rum & Cola / Bar"))
