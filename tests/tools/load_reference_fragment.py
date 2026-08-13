"""Заливка фрагмента референсной книги во временное окружение (план — задача 44, ТЗ 10.1).

Зачем это существует
--------------------

ТЗ 10.1 разрешает ровно одно применение референсной книги: «только для проверки
работоспособности и удобства во временное окружение можно залить фрагмент — чек-лист
открытия на восьми группах и десяток рецептов. Смысл — увидеть, читаемо ли это на телефоне
при настоящем объёме». Двадцать выдуманных пунктов такого ответа не дают: чек-лист бара
Invasion — это восемь групп и три десятка пунктов, часть из которых длиннее строки экрана.

Три режима::

    python -m tests.tools.load_reference_fragment              # залить фрагмент
    python -m tests.tools.load_reference_fragment --purge       # убрать ровно его же
    python -m tests.tools.load_reference_fragment --measure     # медиана и p95 (ТЗ 9)

Границы, которые нельзя переходить
----------------------------------

* **Не поставка.** Модуль лежит под ``tests/`` (решение D13), в образ не попадает
  (``.dockerignore``) и не может быть импортирован из ``src/`` — это проверяет
  ``tests/test_no_seed_data.py::test_src_does_not_import_tests`` разбором AST.
* **Падает при ``APP_ENV=prod``.** Проверка стоит и в ``main()``, и в каждой точке входа,
  которая открывает транзакцию: единственный способ вызвать заливку в проде — обойти обе.
* **Ни одной строки данных в самом файле.** Тексты пунктов и названия рецептов приходят
  из книги; здесь есть только имена листов и заголовки колонок, то есть структура.

Строки-продолжения — правило, которое понадобится и парсеру этапа 1
-------------------------------------------------------------------

На листе «Открытиезакрытие бара» пункт живёт не всегда в одной строке. Пункт «Заполненный
шкаф со всеми печеньями,» продолжается на следующей строке — «шоколадками, гарнировками,
чипсами, прочее...», и наивное правило «каждая непустая ячейка Описание — отдельный пункт»
даёт бармену обрывок фразы и лишнюю галочку.

Признаков продолжения два, и достаточно любого (:func:`is_continuation`):

* предыдущая строка кончается на запятую, точку с запятой, ``+``, ``/`` или дефис —
  человек, который прервал фразу, оставляет след пунктуацией;
* текущая строка начинается со строчной буквы — предложение так не начинают.

Оба признака в референсе совпадают на одном и том же пункте, и оба ложно не срабатывают ни
на одном из остальных: все прочие пункты начинаются с прописной буквы и кончаются буквой,
скобкой или восклицательным знаком. Правило намеренно консервативно — при сомнении строка
считается новым пунктом, потому что лишняя галочка чинится редактором чек-листа за секунду,
а потерянный пункт не заметит никто.

У рецептов такого правила не нужно вовсе, и это важное различие: там блок рецепта
ограничен строками-заголовками, поэтому многострочный гарнир («Огуречный пил» + «с
фигурными» + «дырами ( 20 грм )») склеивается по границе блока, без догадок.

Что разбирается, а что нет
--------------------------

Блок «Открытие» — да. Блок «Закрытие» на том же листе — нет: в его колонке ``№`` посреди
таблицы стоят подзаголовки («Гаджеты на зарядке:», «Наличие:»), то есть у него другая
структура, и разбирать её вслепую ради этапа 0 нечего — чек-лист закрытия по плану уезжает
на этап 1. Пусть лучше не будет кода, чем будет наполовину верный.

Количества не нормализуются (решение D4, ТЗ 7)
----------------------------------------------

Числовая ячейка становится ``qty``, текстовая — ``qty_text`` дословно: «60 мл», «топ»,
«125 мл ( топ )», «3 дропс». В книге нет отдельной колонки единиц, поэтому ``unit``
остаётся пустым всегда, и ветка карточки «qty + unit» этим фрагментом не проверяется —
проверяются две другие ветки задачи 18.
"""

from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import math
import os
import re
import statistics
import time
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import CursorResult, Result, delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from src.config import settings
from src.db.models import (
    ChecklistItem,
    ChecklistRun,
    ChecklistRunItem,
    ChecklistTemplate,
    ChecklistType,
    Recipe,
    RecipeIngredient,
    User,
    Venue,
    VenueMember,
    VenueSettings,
)

# --------------------------------------------------------------------------------------
# Где книга и что в ней искать
# --------------------------------------------------------------------------------------

REPO_ROOT: Final = Path(__file__).resolve().parents[2]

#: Обезличенная копия рабочей книги бара Invasion Universe (см. sanitize_reference.py).
DEFAULT_WORKBOOK: Final = REPO_ROOT / "tests" / "fixtures" / "reference" / "Invasion.xlsx"

#: Лист с чек-листами открытия и закрытия. Имя записано так же, как в книге.
CHECKLIST_SHEET: Final = "Открытиезакрытие бара"

#: Заголовки колонок блока чек-листа, в нормализованном виде.
CHECKLIST_NUMBER_HEADER: Final = "№"
CHECKLIST_GROUP_HEADER: Final = "название:"
CHECKLIST_ITEM_HEADER: Final = "описание:"

#: Листы с рецептами и порядок, в котором из них берётся фрагмент. Два листа взяты
#: намеренно: у «Классики» нет колонки «Лёд» и количества записаны числами, у «Авторки»
#: колонка «Лёд» есть и количества текстовые («60 мл», «топ»). Одного листа мало, чтобы
#: увидеть обе формы на экране телефона.
RECIPE_SHEETS: Final = ("Классика", "Авторка")

#: Заголовки колонок рецепта. Написания перечислены вместе с опечатками книги
#: («Ингридиенты» на листе «Авторка»): парсер обязан находить колонку, а не спорить с ней.
RECIPE_NAME_HEADERS: Final = frozenset({"название"})
RECIPE_INGREDIENT_HEADERS: Final = frozenset(
    {"ингредиент", "ингредиенты", "ингридиент", "ингридиенты"}
)
RECIPE_QTY_HEADERS: Final = frozenset({"кол-во", "количество"})
RECIPE_GARNISH_HEADERS: Final = frozenset({"гарнир"})
RECIPE_GLASSWARE_HEADERS: Final = frozenset({"посуда"})
RECIPE_ICE_HEADERS: Final = frozenset({"лёд", "лед"})
RECIPE_METHOD_HEADERS: Final = frozenset({"метод"})

#: «Десяток рецептов» из ТЗ 10.1, поровну с каждого листа.
DEFAULT_RECIPE_LIMIT: Final = 12

#: Сколько строк от начала листа просматривается в поисках строки заголовков.
HEADER_SEARCH_DEPTH: Final = 10

# --------------------------------------------------------------------------------------
# Временное заведение, если своё не передали
# --------------------------------------------------------------------------------------

#: Имя служебного заведения, которое создаётся при запуске без ``--venue-id``. Латиница и
#: слово «fragment» — чтобы его нельзя было спутать с настоящим заведением заказчика.
SCRATCH_VENUE_NAME: Final = "Reference fragment"
SCRATCH_VENUE_CITY: Final = "Reference"
SCRATCH_VENUE_TIMEZONE: Final = "Europe/Moscow"
SCRATCH_SHIFT_START: Final = dt.time(8, 0)
SCRATCH_SHIFT_END: Final = dt.time(23, 0)

#: Имя шаблона чек-листа, если его пришлось создать. Мастер заведения (задача 26) создаёт
#: свой; здесь имя нужно только потому, что колонка NOT NULL.
FRAGMENT_TEMPLATE_NAME: Final = "opening"

#: telegram_id технического пользователя режима замеров. Отрицательное значение Telegram
#: не выдаёт никогда, поэтому столкнуться с живым аккаунтом невозможно. Строка живёт
#: внутри транзакции, которую режим замеров всегда откатывает.
MEASUREMENT_TELEGRAM_ID: Final = -1
MEASUREMENT_USER_NAME: Final = "measurement"

# --------------------------------------------------------------------------------------
# Замеры (ТЗ 9)
# --------------------------------------------------------------------------------------

#: ТЗ 9: «время ответа на нажатие кнопки — до 1 секунды, поиск по рецептам — до 2 секунд».
CHECKLIST_BUDGET_MS: Final = 1000.0
SEARCH_BUDGET_MS: Final = 2000.0

#: Порог trigram-похожести, с которым будет работать поиск задачи 13.
SIMILARITY_THRESHOLD: Final = 0.3

#: Страница поиска: десять по ТЗ 5.5 плюс одна, чтобы понять, есть ли следующая.
SEARCH_PAGE: Final = 11

DEFAULT_REPEAT: Final = 30

#: Лимит одного сообщения Telegram. Чек-лист рисуется текстом (решение D11), и на реальном
#: объёме в него нужно попасть — это ровно тот вопрос, ради которого фрагмент и заливается.
TELEGRAM_MESSAGE_LIMIT: Final = 4096

# --------------------------------------------------------------------------------------
# Правило склейки строк-продолжений
# --------------------------------------------------------------------------------------

#: Знаки, на которых обрывается незаконченная фраза.
OPEN_ENDINGS: Final = (",", ";", "+", "/", "-", "–", "—")

#: Дефис в конце — это перенос слова, склейка идёт без пробела.
HYPHEN_ENDINGS: Final = ("-", "–", "—")

_WHITESPACE: Final = re.compile(r"\s+")


class ReferenceWorkbookError(RuntimeError):
    """Книга не той структуры: нет листа, нет колонки, нет строки заголовков."""


class ProductionRefusedError(RuntimeError):
    """Попытка запустить инструмент в проде (решение D13)."""

    def __init__(self) -> None:
        super().__init__(
            "APP_ENV=prod: загрузчик фрагмента референсной книги пишет данные в базу и в "
            "продовое окружение не допускается (ТЗ 10.1, решение D13)"
        )


class VenueNotFoundError(RuntimeError):
    """Заведения с таким id нет — заливать фрагмент некуда."""

    def __init__(self, venue_id: int) -> None:
        super().__init__(f"заведение {venue_id} не найдено")
        self.venue_id = venue_id


def is_production() -> bool:
    """Прод ли это. Смотрит и на окружение, и на прочитанные настройки.

    Две проверки вместо одной: ``settings`` — синглтон, собранный при импорте, поэтому
    переменная, выставленная позже, в него не попадёт; а ``.env`` с ``APP_ENV=prod``
    не попадёт в ``os.environ``. Прод — если так считает хотя бы один из двух источников.
    """
    if os.environ.get("APP_ENV", "").strip().lower() == "prod":
        return True
    return settings.is_production


def guard_not_production() -> None:
    """Единственная строчка, которую обязана звать каждая точка входа с записью в базу."""
    if is_production():
        raise ProductionRefusedError


# --------------------------------------------------------------------------------------
# Разобранные структуры (никакой базы — это половина разбирается и тестируется отдельно)
# --------------------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class ParsedGroup:
    """Группа чек-листа: имя из колонки «Название:» и пункты из «Описание:» (ТЗ 5.4)."""

    index: int
    name: str
    items: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class ParsedIngredient:
    """Строка состава. ``qty`` — только для по-настоящему числовой ячейки (решение D4)."""

    name: str
    qty: Decimal | None
    qty_text: str | None
    order_index: int


@dataclass(frozen=True, slots=True)
class ParsedRecipe:
    """Рецепт как он лежит в книге. Категория — имя листа, других категорий там нет."""

    name: str
    category: str
    glassware: str | None
    method: str | None
    ice: str | None
    garnish: str | None
    ingredients: tuple[ParsedIngredient, ...]


@dataclass(frozen=True, slots=True)
class Fragment:
    """Всё, что заливается и удаляется одной командой."""

    groups: tuple[ParsedGroup, ...]
    recipes: tuple[ParsedRecipe, ...]

    @property
    def item_count(self) -> int:
        return sum(len(group.items) for group in self.groups)

    @property
    def item_texts(self) -> tuple[str, ...]:
        return tuple(text for group in self.groups for text in group.items)


# --------------------------------------------------------------------------------------
# Чтение книги
# --------------------------------------------------------------------------------------


def normalise(value: object) -> str:
    """Заголовок или текст ячейки без регистра, переносов и двойных пробелов."""
    return collapse(value).lower()


def collapse(value: object) -> str:
    """Текст ячейки в одну строку. Число, ставшее ``45.0``, снова становится ``45``."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _WHITESPACE.sub(" ", str(value)).strip()


def numeric(value: object) -> Decimal | None:
    """``Decimal`` для числовой ячейки и ``None`` для всего остального.

    Строка «60 мл» числом не считается намеренно: единицы не нормализуются (задача 44),
    такая ячейка целиком уходит в ``qty_text``.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    return None


def open_workbook(path: Path) -> Workbook:
    if not path.exists():
        raise ReferenceWorkbookError(f"референсной книги нет: {path}")
    return load_workbook(path, data_only=True, read_only=True)


def sheet(workbook: Workbook, name: str) -> Worksheet:
    if name not in workbook.sheetnames:
        raise ReferenceWorkbookError(f"в книге нет листа {name!r}")
    found = workbook[name]
    if not isinstance(found, Worksheet):
        raise ReferenceWorkbookError(f"лист {name!r} не таблица")
    return found


def rows_of(workbook: Workbook, name: str) -> list[tuple[object, ...]]:
    """Лист как список строк. Книга маленькая, и работать со списком проще, чем с курсором."""
    return [tuple(row) for row in sheet(workbook, name).iter_rows(values_only=True)]


def cell(row: Sequence[object], index: int | None) -> object:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def text_at(row: Sequence[object], index: int | None) -> str:
    return collapse(cell(row, index))


# --------------------------------------------------------------------------------------
# Строки-продолжения
# --------------------------------------------------------------------------------------


def is_continuation(previous: str | None, text: str) -> bool:
    """Продолжение ли ``text`` предыдущего пункта. Разбор правила — в докстроке модуля."""
    if not previous or not text:
        return False
    if previous.endswith(OPEN_ENDINGS):
        return True
    first = text[0]
    return first.isalpha() and first.islower()


def join_continuation(previous: str, addition: str) -> str:
    """Склейка: через пробел, а после дефиса — без него, потому что это перенос слова."""
    if previous.endswith(HYPHEN_ENDINGS):
        return previous[:-1] + addition
    return f"{previous} {addition}"


def merge_continuations(lines: Iterable[str]) -> tuple[str, ...]:
    """Последовательность ячеек «Описание» -> пункты чек-листа."""
    items: list[str] = []
    for line in lines:
        text = collapse(line)
        if not text:
            continue
        if items and is_continuation(items[-1], text):
            items[-1] = join_continuation(items[-1], text)
        else:
            items.append(text)
    return tuple(items)


# --------------------------------------------------------------------------------------
# Разбор чек-листа
# --------------------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class ChecklistColumns:
    """Где на листе лежит блок чек-листа. Индексы нулевые."""

    header_row: int
    number: int
    group: int
    item: int


def find_checklist_columns(rows: Sequence[tuple[object, ...]]) -> ChecklistColumns:
    """Первый (то есть левый, «Открытие») блок листа.

    Колонки ищутся по заголовкам, а не по буквам: блоков на листе два, и второй начинается
    там, где кончается первый, — привязка к ``A/B/C`` рассыплется от одной вставленной
    колонки.
    """
    for index, row in enumerate(rows[:HEADER_SEARCH_DEPTH]):
        headers = [normalise(value) for value in row]
        if CHECKLIST_ITEM_HEADER not in headers or CHECKLIST_GROUP_HEADER not in headers:
            continue
        number = headers.index(CHECKLIST_NUMBER_HEADER) if CHECKLIST_NUMBER_HEADER in headers else 0
        return ChecklistColumns(
            header_row=index,
            number=number,
            group=headers.index(CHECKLIST_GROUP_HEADER),
            item=headers.index(CHECKLIST_ITEM_HEADER),
        )
    raise ReferenceWorkbookError(
        f"на листе {CHECKLIST_SHEET!r} не найдена строка заголовков "
        f"({CHECKLIST_GROUP_HEADER!r} и {CHECKLIST_ITEM_HEADER!r})"
    )


def parse_opening_checklist(workbook: Workbook) -> tuple[ParsedGroup, ...]:
    """Чек-лист открытия: группы в порядке листа, пункты со склейкой продолжений.

    Группа без пунктов не создаётся — решение D2: «группа не существует без пунктов».
    """
    rows = rows_of(workbook, CHECKLIST_SHEET)
    columns = find_checklist_columns(rows)

    names: list[str] = []
    buckets: list[list[str]] = []

    for row in rows[columns.header_row + 1 :]:
        group_name = text_at(row, columns.group)
        item_text = text_at(row, columns.item)

        if group_name:
            names.append(group_name)
            buckets.append([])
        if not item_text or not buckets:
            # Пункт до первой группы игнорируется: у него нет заголовка, под которым его
            # показать, и в референсе такого нет.
            continue

        current = buckets[-1]
        if current and is_continuation(current[-1], item_text):
            current[-1] = join_continuation(current[-1], item_text)
        else:
            current.append(item_text)

    groups = [
        ParsedGroup(index=index, name=name, items=tuple(items))
        for index, (name, items) in enumerate(zip(names, buckets, strict=True))
        if items
    ]
    if not groups:
        raise ReferenceWorkbookError(f"на листе {CHECKLIST_SHEET!r} не нашлось ни одной группы")
    return tuple(groups)


# --------------------------------------------------------------------------------------
# Разбор рецептов
# --------------------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class RecipeColumns:
    """Раскладка колонок листа рецептов. ``ice`` есть не на всех листах."""

    header_row: int
    name: int
    ingredient: int
    qty: int | None
    garnish: int | None
    glassware: int | None
    ice: int | None
    method: int | None


def _column(headers: Sequence[str], names: frozenset[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in names:
            return index
    return None


def find_recipe_columns(rows: Sequence[tuple[object, ...]], sheet_name: str) -> RecipeColumns:
    for index, row in enumerate(rows[:HEADER_SEARCH_DEPTH]):
        headers = [normalise(value) for value in row]
        name = _column(headers, RECIPE_NAME_HEADERS)
        ingredient = _column(headers, RECIPE_INGREDIENT_HEADERS)
        if name is None or ingredient is None:
            continue
        return RecipeColumns(
            header_row=index,
            name=name,
            ingredient=ingredient,
            qty=_column(headers, RECIPE_QTY_HEADERS),
            garnish=_column(headers, RECIPE_GARNISH_HEADERS),
            glassware=_column(headers, RECIPE_GLASSWARE_HEADERS),
            ice=_column(headers, RECIPE_ICE_HEADERS),
            method=_column(headers, RECIPE_METHOD_HEADERS),
        )
    raise ReferenceWorkbookError(f"на листе {sheet_name!r} не найдена строка заголовков рецепта")


def _is_header_row(row: Sequence[object], columns: RecipeColumns) -> bool:
    return normalise(cell(row, columns.name)) in RECIPE_NAME_HEADERS


def _blocks(
    rows: Sequence[tuple[object, ...]],
    columns: RecipeColumns,
) -> list[list[tuple[object, ...]]]:
    """Разрезать лист на блоки рецептов.

    Блок начинается строкой сразу за строкой-заголовком и кончается следующим заголовком
    или первой строкой, где пусты и название, и ингредиент, — так отсекается мусор в хвосте
    листа, который иначе приклеился бы к гарниру последнего рецепта.
    """
    blocks: list[list[tuple[object, ...]]] = []
    current: list[tuple[object, ...]] | None = None

    for row in rows[columns.header_row :]:
        if _is_header_row(row, columns):
            current = []
            blocks.append(current)
            continue
        if current is None:
            continue
        if not text_at(row, columns.name) and not text_at(row, columns.ingredient):
            current = None
            continue
        current.append(row)

    return [block for block in blocks if block]


def _joined(block: Sequence[Sequence[object]], index: int | None) -> str | None:
    """Значение колонки, размазанное по строкам блока, одной строкой.

    Здесь эвристика продолжений не нужна и не применяется: границы блока уже сказали, какие
    строки относятся к одному рецепту, — «Огуречный пил» + «с фигурными» + «дырами ( 20 грм )»
    склеиваются по построению.
    """
    if index is None:
        return None
    parts = [text_at(row, index) for row in block]
    joined = " ".join(part for part in parts if part)
    return joined or None


def _ingredients(
    block: Sequence[Sequence[object]], columns: RecipeColumns
) -> tuple[ParsedIngredient, ...]:
    lines: list[ParsedIngredient] = []
    for row in block:
        name = text_at(row, columns.ingredient)
        if not name:
            continue
        raw = cell(row, columns.qty)
        qty = numeric(raw)
        lines.append(
            ParsedIngredient(
                name=name,
                qty=qty,
                qty_text=None if qty is not None else (collapse(raw) or None),
                order_index=len(lines),
            )
        )
    return tuple(lines)


def parse_recipe_sheet(workbook: Workbook, sheet_name: str) -> tuple[ParsedRecipe, ...]:
    """Все рецепты одного листа. Категория рецепта — имя листа."""
    rows = rows_of(workbook, sheet_name)
    columns = find_recipe_columns(rows, sheet_name)

    recipes: list[ParsedRecipe] = []
    for block in _blocks(rows, columns):
        name = text_at(block[0], columns.name)
        ingredients = _ingredients(block, columns)
        if not name or not ingredients:
            continue
        recipes.append(
            ParsedRecipe(
                name=name,
                category=sheet_name,
                glassware=_joined(block, columns.glassware),
                method=_joined(block, columns.method),
                ice=_joined(block, columns.ice),
                garnish=_joined(block, columns.garnish),
                ingredients=ingredients,
            )
        )
    return tuple(recipes)


def parse_recipes(workbook: Workbook, *, limit: int) -> tuple[ParsedRecipe, ...]:
    """Фрагмент рецептов, поровну с каждого листа из :data:`RECIPE_SHEETS`."""
    if limit <= 0:
        return ()
    quota, extra = divmod(limit, len(RECIPE_SHEETS))
    taken: list[ParsedRecipe] = []
    for position, sheet_name in enumerate(RECIPE_SHEETS):
        share = quota + (1 if position < extra else 0)
        taken.extend(parse_recipe_sheet(workbook, sheet_name)[:share])
    return tuple(taken)


def parse_fragment(
    path: Path = DEFAULT_WORKBOOK,
    *,
    recipe_limit: int = DEFAULT_RECIPE_LIMIT,
) -> Fragment:
    """Вся книга -> структуры. Базы данных этот путь не касается вовсе."""
    workbook = open_workbook(path)
    try:
        return Fragment(
            groups=parse_opening_checklist(workbook),
            recipes=parse_recipes(workbook, limit=recipe_limit),
        )
    finally:
        workbook.close()


# --------------------------------------------------------------------------------------
# Отчёты
# --------------------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class LoadReport:
    venue_id: int
    template_id: int
    groups: int
    items_added: int
    items_skipped: int
    recipes_added: int
    recipes_skipped: int
    ingredients_added: int


@dataclass(frozen=True, slots=True)
class PurgeReport:
    venue_id: int | None
    runs_deleted: int
    items_deleted: int
    recipes_deleted: int


@dataclass(frozen=True, slots=True)
class Timing:
    """Медиана и p95 одного замера против бюджета ТЗ 9."""

    label: str
    samples: int
    median_ms: float
    p95_ms: float
    budget_ms: float

    @property
    def within_budget(self) -> bool:
        return self.p95_ms <= self.budget_ms


@dataclass(frozen=True, slots=True)
class MeasurementReport:
    venue_id: int
    items: int
    recipes: int
    message_chars: int
    timings: tuple[Timing, ...]

    @property
    def fits_one_message(self) -> bool:
        return self.message_chars <= TELEGRAM_MESSAGE_LIMIT

    @property
    def within_budget(self) -> bool:
        return all(timing.within_budget for timing in self.timings)


def percentile(samples: Sequence[float], fraction: float) -> float:
    """Перцентиль методом ближайшего ранга: на 30 замерах интерполяция врёт больше, чем даёт."""
    if not samples:
        return 0.0
    ordered = sorted(samples)
    rank = max(1, math.ceil(fraction * len(ordered)))
    return ordered[min(rank, len(ordered)) - 1]


def summarise(label: str, samples: Sequence[float], budget_ms: float) -> Timing:
    return Timing(
        label=label,
        samples=len(samples),
        median_ms=statistics.median(samples) if samples else 0.0,
        p95_ms=percentile(samples, 0.95),
        budget_ms=budget_ms,
    )


# --------------------------------------------------------------------------------------
# База: куда заливать
# --------------------------------------------------------------------------------------


async def resolve_venue(session: AsyncSession, venue_id: int | None) -> int:
    """Заведение фрагмента: переданное, ранее созданное служебное или новое служебное.

    Основной путь — ``--venue-id``: по ТЗ 10.1 тестовое заведение поднимается через
    интерфейс, и фрагмент льётся именно в него. Служебное заведение — запасной путь, чтобы
    инструмент был самодостаточен на пустой базе.
    """
    if venue_id is not None:
        found = await session.scalar(select(Venue.id).where(Venue.id == venue_id))
        if found is None:
            raise VenueNotFoundError(venue_id)
        return int(found)

    existing = await session.scalar(select(Venue.id).where(Venue.name == SCRATCH_VENUE_NAME))
    if existing is not None:
        return int(existing)

    venue = Venue(
        name=SCRATCH_VENUE_NAME,
        city=SCRATCH_VENUE_CITY,
        timezone=SCRATCH_VENUE_TIMEZONE,
    )
    session.add(venue)
    await session.flush()
    session.add(
        VenueSettings(
            venue_id=venue.id,
            default_shift_start=SCRATCH_SHIFT_START,
            default_shift_end=SCRATCH_SHIFT_END,
        )
    )
    await session.flush()
    return venue.id


async def find_venue(session: AsyncSession, venue_id: int | None) -> int | None:
    """То же, что :func:`resolve_venue`, но ничего не создаёт: путь ``--purge``."""
    if venue_id is not None:
        found = await session.scalar(select(Venue.id).where(Venue.id == venue_id))
        return int(found) if found is not None else None
    existing = await session.scalar(select(Venue.id).where(Venue.name == SCRATCH_VENUE_NAME))
    return int(existing) if existing is not None else None


async def opening_template(session: AsyncSession, venue_id: int) -> ChecklistTemplate:
    """Активный шаблон открытия заведения; создаётся, только если его ещё нет.

    Мастер создания заведения (задача 26) заводит пустые шаблоны сам (решение B1), поэтому
    штатно эта функция ничего не создаёт, а находит.
    """
    template = await session.scalar(
        select(ChecklistTemplate).where(
            ChecklistTemplate.venue_id == venue_id,
            ChecklistTemplate.type == ChecklistType.OPENING,
            ChecklistTemplate.is_active.is_(True),
        )
    )
    if template is not None:
        return template

    template = ChecklistTemplate(
        venue_id=venue_id,
        type=ChecklistType.OPENING,
        name=FRAGMENT_TEMPLATE_NAME,
        version=1,
        is_active=True,
    )
    session.add(template)
    await session.flush()
    return template


# --------------------------------------------------------------------------------------
# Заливка
# --------------------------------------------------------------------------------------


async def load_fragment(
    session: AsyncSession,
    *,
    venue_id: int | None = None,
    fragment: Fragment | None = None,
) -> LoadReport:
    """Залить фрагмент. Повторный запуск ничего не задваивает.

    Идемпотентность не роскошь: у ``recipes`` есть уникальный ключ
    ``(venue_id, category, lower(btrim(name)))`` (решение D6), и второй запуск без проверки
    просто упал бы посреди заливки, оставив половину.
    """
    guard_not_production()
    parsed = fragment if fragment is not None else parse_fragment()
    resolved = await resolve_venue(session, venue_id)
    template = await opening_template(session, resolved)

    items_added, items_skipped = await _load_items(session, template.id, parsed)
    recipes_added, recipes_skipped, ingredients = await _load_recipes(session, resolved, parsed)
    await session.flush()

    return LoadReport(
        venue_id=resolved,
        template_id=template.id,
        groups=len(parsed.groups),
        items_added=items_added,
        items_skipped=items_skipped,
        recipes_added=recipes_added,
        recipes_skipped=recipes_skipped,
        ingredients_added=ingredients,
    )


async def _load_items(
    session: AsyncSession,
    template_id: int,
    fragment: Fragment,
) -> tuple[int, int]:
    present = set(
        (
            await session.scalars(
                select(ChecklistItem.text).where(ChecklistItem.template_id == template_id)
            )
        ).all()
    )

    added = 0
    skipped = 0
    order = 0
    for group in fragment.groups:
        for text in group.items:
            order += 1
            if text in present:
                skipped += 1
                continue
            session.add(
                ChecklistItem(
                    template_id=template_id,
                    group_name=group.name,
                    group_index=group.index,
                    order_index=order,
                    text=text,
                )
            )
            added += 1
    return added, skipped


def _recipe_key(name: str, category: str) -> tuple[str, str]:
    """Ключ уникальности рецепта из решения D6, посчитанный на стороне Python."""
    return category.strip(), name.strip().lower()


async def _load_recipes(
    session: AsyncSession,
    venue_id: int,
    fragment: Fragment,
) -> tuple[int, int, int]:
    rows = (
        await session.execute(
            select(Recipe.name, Recipe.category).where(Recipe.venue_id == venue_id)
        )
    ).all()
    present = {_recipe_key(name, category) for name, category in rows}

    added = 0
    skipped = 0
    ingredients = 0
    for parsed in fragment.recipes:
        key = _recipe_key(parsed.name, parsed.category)
        if key in present:
            skipped += 1
            continue
        present.add(key)
        recipe = Recipe(
            venue_id=venue_id,
            name=parsed.name,
            category=parsed.category,
            glassware=parsed.glassware,
            method=parsed.method,
            ice=parsed.ice,
            garnish=parsed.garnish,
        )
        session.add(recipe)
        await session.flush()
        for line in parsed.ingredients:
            session.add(
                RecipeIngredient(
                    recipe_id=recipe.id,
                    name=line.name,
                    qty=line.qty,
                    qty_text=line.qty_text,
                    order_index=line.order_index,
                )
            )
            ingredients += 1
        added += 1
    return added, skipped, ingredients


# --------------------------------------------------------------------------------------
# Удаление
# --------------------------------------------------------------------------------------


def _rows_affected(result: Result[Any]) -> int:
    """How many rows a `DELETE` actually removed.

    `AsyncSession.execute` is declared as returning `Result`, which has no `rowcount` —
    only `CursorResult`, what a DML statement returns at run time, does. The narrowing is a
    real check rather than a cast, so a driver that ever hands back something else reports
    nothing removed instead of raising. DBAPIs also use -1 for "unknown", which is not a
    count and must not be added into one.
    """
    if not isinstance(result, CursorResult):
        return 0
    return max(result.rowcount, 0)


async def purge_fragment(
    session: AsyncSession,
    *,
    venue_id: int | None = None,
    fragment: Fragment | None = None,
) -> PurgeReport:
    """Убрать ровно то, что залил бы :func:`load_fragment`, и ничего сверх того.

    Удаление идёт не «по заведению», а по свежему разбору книги: то, что заведение завело
    руками, останется на месте, даже если лежит в том же шаблоне и в той же категории.

    Прогоны чек-листа, которые ссылаются на пункты фрагмента, удаляются вместе с ними:
    ``checklist_run_items.item_id`` — это RESTRICT, и прогон по данным фрагмента сам такие
    же временные данные. Само заведение и пустой шаблон остаются: пустой шаблон — штатное
    состояние заведения (решение B1), а не мусор.
    """
    guard_not_production()
    parsed = fragment if fragment is not None else parse_fragment()
    resolved = await find_venue(session, venue_id)
    if resolved is None:
        return PurgeReport(venue_id=None, runs_deleted=0, items_deleted=0, recipes_deleted=0)

    templates = list(
        (
            await session.scalars(
                select(ChecklistTemplate.id).where(ChecklistTemplate.venue_id == resolved)
            )
        ).all()
    )

    item_ids: list[int] = []
    if templates:
        item_ids = list(
            (
                await session.scalars(
                    select(ChecklistItem.id).where(
                        ChecklistItem.template_id.in_(templates),
                        ChecklistItem.text.in_(parsed.item_texts),
                    )
                )
            ).all()
        )

    runs_deleted = 0
    items_deleted = 0
    if item_ids:
        run_ids = list(
            (
                await session.scalars(
                    select(ChecklistRunItem.run_id)
                    .where(ChecklistRunItem.item_id.in_(item_ids))
                    .distinct()
                )
            ).all()
        )
        if run_ids:
            deleted_runs = await session.execute(
                delete(ChecklistRun).where(
                    ChecklistRun.venue_id == resolved,
                    ChecklistRun.id.in_(run_ids),
                )
            )
            runs_deleted = _rows_affected(deleted_runs)
        deleted_items = await session.execute(
            delete(ChecklistItem).where(ChecklistItem.id.in_(item_ids))
        )
        items_deleted = _rows_affected(deleted_items)

    recipes_deleted = 0
    for parsed_recipe in parsed.recipes:
        deleted = await session.execute(
            delete(Recipe).where(
                Recipe.venue_id == resolved,
                Recipe.category == parsed_recipe.category,
                func.lower(func.btrim(Recipe.name)) == parsed_recipe.name.strip().lower(),
            )
        )
        recipes_deleted += _rows_affected(deleted)

    await session.flush()
    return PurgeReport(
        venue_id=resolved,
        runs_deleted=runs_deleted,
        items_deleted=items_deleted,
        recipes_deleted=recipes_deleted,
    )


# --------------------------------------------------------------------------------------
# Замеры (ТЗ 9: кнопка до 1 секунды, поиск до 2 секунд)
# --------------------------------------------------------------------------------------


def render_checklist(
    items: Sequence[ChecklistItem],
    done: frozenset[int],
) -> str:
    """Текст сообщения чек-листа — то, что задача 24 будет рисовать по-настоящему.

    Формулировки задаче 21, а не сюда; здесь нужен только объём: сколько символов занимает
    восемь групп и три десятка пунктов и влезает ли это в одно сообщение (решение D11).
    """
    lines: list[str] = []
    group: str | None = None
    for item in sorted(items, key=lambda row: (row.group_index, row.order_index, row.id)):
        if item.group_name != group:
            group = item.group_name
            lines.append("")
            lines.append(group or "")
        lines.append(f"[{'x' if item.id in done else ' '}] {item.text}")
    return "\n".join(lines).strip()


def search_probes(names: Sequence[str]) -> tuple[str, ...]:
    """Запросы для замера: как есть, регистром и с опечаткой (ТЗ 5.5, тест 40).

    Строки берутся из самого фрагмента, а не пишутся здесь: в этом файле не должно быть
    ни одного названия из книги.
    """
    probes: list[str] = []
    for name in names:
        cleaned = name.strip()
        if len(cleaned) < 4:
            continue
        probes.append(cleaned)
        probes.append(cleaned.upper())
        probes.append(cleaned.lower())
        # Выпавшая буква — самая частая опечатка на телефоне одной рукой.
        probes.append(cleaned[: len(cleaned) // 2] + cleaned[len(cleaned) // 2 + 1 :])
    return tuple(probes)


async def _search_once(session: AsyncSession, venue_id: int, query: str) -> int:
    """Поиск в той форме, в которой его сделает задача 13: ILIKE плюс trigram-похожесть."""
    statement = (
        select(Recipe.id)
        .where(
            or_(Recipe.venue_id == venue_id, Recipe.venue_id.is_(None)),
            Recipe.is_active.is_(True),
            or_(
                Recipe.name.ilike(f"%{query}%"),
                func.similarity(Recipe.name, query) > SIMILARITY_THRESHOLD,
            ),
        )
        .order_by(func.similarity(Recipe.name, query).desc(), Recipe.name)
        .limit(SEARCH_PAGE)
    )
    return len((await session.scalars(statement)).all())


async def _measurement_user(session: AsyncSession, venue_id: int) -> int:
    """Кто «проходит» чек-лист в замере: любой участник заведения, иначе технический."""
    member = await session.scalar(
        select(VenueMember.user_id).where(VenueMember.venue_id == venue_id).limit(1)
    )
    if member is not None:
        return int(member)

    existing = await session.scalar(
        select(User.id).where(User.telegram_id == MEASUREMENT_TELEGRAM_ID)
    )
    if existing is not None:
        return int(existing)

    user = User(telegram_id=MEASUREMENT_TELEGRAM_ID, full_name=MEASUREMENT_USER_NAME)
    session.add(user)
    await session.flush()
    return user.id


async def measure(
    session: AsyncSession,
    *,
    venue_id: int | None = None,
    repeat: int = DEFAULT_REPEAT,
) -> MeasurementReport:
    """Медиана и p95 для отрисовки чек-листа и для поиска рецепта.

    Вызывающий обязан откатить транзакцию: замер создаёт прогон чек-листа, чтобы мерить
    тот же путь, которым пойдёт задача 17 (пункты шаблона плюс состояние прогона), и
    оставлять этот прогон в базе незачем. И ``main()``, и db-тест это делают.
    """
    guard_not_production()
    resolved = await find_venue(session, venue_id)
    if resolved is None:
        raise VenueNotFoundError(venue_id if venue_id is not None else 0)

    template = await opening_template(session, resolved)
    items = list(
        (
            await session.scalars(
                select(ChecklistItem).where(ChecklistItem.template_id == template.id)
            )
        ).all()
    )
    names = list(
        (await session.scalars(select(Recipe.name).where(Recipe.venue_id == resolved))).all()
    )

    run = ChecklistRun(
        venue_id=resolved,
        shift_id=None,
        user_id=await _measurement_user(session, resolved),
        template_id=template.id,
        type=ChecklistType.OPENING,
        total_items=len(items),
        sent_at=dt.datetime.now(dt.UTC),
    )
    session.add(run)
    await session.flush()
    for item in items:
        session.add(ChecklistRunItem(run_id=run.id, item_id=item.id))
    await session.flush()

    run_id = run.id
    template_id = template.id
    message = ""
    render_samples: list[float] = []
    for _ in range(max(repeat, 1)):
        # Кэш сессии выключается руками: иначе второй проход собирает объекты из identity
        # map и меряет скорость словаря, а не базы.
        session.expunge_all()
        started = time.perf_counter()
        rows = list(
            (
                await session.scalars(
                    select(ChecklistItem).where(ChecklistItem.template_id == template_id)
                )
            ).all()
        )
        state = (
            await session.scalars(
                select(ChecklistRunItem.item_id).where(
                    ChecklistRunItem.run_id == run_id,
                    ChecklistRunItem.is_done.is_(True),
                )
            )
        ).all()
        message = render_checklist(rows, frozenset(state))
        render_samples.append((time.perf_counter() - started) * 1000)

    probes = search_probes(names)
    search_samples: list[float] = []
    for index in range(max(repeat, 1)):
        if not probes:
            break
        session.expunge_all()
        started = time.perf_counter()
        await _search_once(session, resolved, probes[index % len(probes)])
        search_samples.append((time.perf_counter() - started) * 1000)

    timings = [summarise("checklist", render_samples, CHECKLIST_BUDGET_MS)]
    if search_samples:
        timings.append(summarise("search", search_samples, SEARCH_BUDGET_MS))

    return MeasurementReport(
        venue_id=resolved,
        items=len(items),
        recipes=len(names),
        message_chars=len(message),
        timings=tuple(timings),
    )


# --------------------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="load_reference_fragment",
        description=(
            "Фрагмент референсной книги во временное окружение: чек-лист открытия на "
            "восьми группах и десяток рецептов (ТЗ 10.1). В прод не допускается."
        ),
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--purge", action="store_true", help="удалить ранее залитый фрагмент")
    mode.add_argument(
        "--measure",
        action="store_true",
        help="замерить медиану и p95 отрисовки чек-листа и поиска рецепта",
    )
    parser.add_argument(
        "--venue-id",
        type=int,
        default=None,
        help="заведение; без него берётся (и при заливке создаётся) служебное",
    )
    parser.add_argument(
        "--recipes",
        type=int,
        default=DEFAULT_RECIPE_LIMIT,
        help=f"сколько рецептов залить (по умолчанию {DEFAULT_RECIPE_LIMIT})",
    )
    parser.add_argument(
        "--repeat",
        type=int,
        default=DEFAULT_REPEAT,
        help=f"сколько раз повторить каждый замер (по умолчанию {DEFAULT_REPEAT})",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="путь к референсной книге",
    )
    return parser


def _print_load(report: LoadReport) -> None:
    print(f"  заведение:          {report.venue_id}")
    print(f"  шаблон открытия:    {report.template_id}")
    print(f"  групп:              {report.groups}")
    print(f"  пунктов добавлено:  {report.items_added} (уже было: {report.items_skipped})")
    print(f"  рецептов добавлено: {report.recipes_added} (уже было: {report.recipes_skipped})")
    print(f"  строк состава:      {report.ingredients_added}")


def _print_purge(report: PurgeReport) -> None:
    if report.venue_id is None:
        print("  заведения с фрагментом не найдено, удалять нечего")
        return
    print(f"  заведение:          {report.venue_id}")
    print(f"  прогонов удалено:   {report.runs_deleted}")
    print(f"  пунктов удалено:    {report.items_deleted}")
    print(f"  рецептов удалено:   {report.recipes_deleted}")


def _print_measure(report: MeasurementReport) -> None:
    print(f"  заведение:          {report.venue_id}")
    print(f"  пунктов / рецептов: {report.items} / {report.recipes}")
    fits = "влезает" if report.fits_one_message else "НЕ ВЛЕЗАЕТ"
    print(f"  сообщение:          {report.message_chars} символов, {fits} в одно сообщение")
    for timing in report.timings:
        verdict = "ok" if timing.within_budget else "ВЫШЕ БЮДЖЕТА"
        print(
            f"  {timing.label:<10} медиана {timing.median_ms:7.1f} мс · "
            f"p95 {timing.p95_ms:7.1f} мс · бюджет {timing.budget_ms:.0f} мс · {verdict}"
        )


async def run(arguments: argparse.Namespace) -> int:
    """Точка входа CLI. Заливка коммитится, замер откатывается всегда."""
    from src.db.session import dispose_engine, get_sessionmaker

    guard_not_production()
    fragment = parse_fragment(arguments.workbook, recipe_limit=arguments.recipes)
    factory = get_sessionmaker()
    try:
        if arguments.measure:
            async with factory() as session:
                try:
                    _print_measure(
                        await measure(
                            session,
                            venue_id=arguments.venue_id,
                            repeat=arguments.repeat,
                        )
                    )
                finally:
                    await session.rollback()
            return 0

        async with factory() as session:
            if arguments.purge:
                _print_purge(
                    await purge_fragment(session, venue_id=arguments.venue_id, fragment=fragment)
                )
            else:
                _print_load(
                    await load_fragment(session, venue_id=arguments.venue_id, fragment=fragment)
                )
            await session.commit()
        return 0
    finally:
        await dispose_engine()


def main(argv: Sequence[str] | None = None) -> int:
    arguments = build_parser().parse_args(argv)
    guard_not_production()
    return asyncio.run(run(arguments))


if __name__ == "__main__":
    raise SystemExit(main())
