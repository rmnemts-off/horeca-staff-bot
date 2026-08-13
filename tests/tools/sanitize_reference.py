"""Обезличивание референсной книги перед коммитом в публичный репозиторий.

Зачем. ТЗ 1.6 требует держать референсную книгу в ``tests/fixtures/reference/`` как эталон
формата для парсеров. Но репозиторий публичный, а исходная книга несёт данные, которые
нельзя публиковать: имена и мобильные телефоны контактных лиц поставщиков (персональные
данные третьих лиц) и коммерческие цифры бара — закупочные цены, себестоимость, FC %,
ставку и суммы.

Что делает скрипт. Читает исходник, заменяет ровно эти значения и пишет обезличенную копию.
Всё, ради чего книга нужна парсерам, сохраняется дословно: имена листов, заголовки колонок
вместе с опечатками, структура блоков, повторяющиеся строки-заголовки, многострочные
значения, нечисловые количества и ячейки, которые Excel превратил в даты.

Побочный эффект пересохранения. openpyxl не считает формулы: при сохранении книги
кэшированные результаты формул теряются, и такие ячейки читаются как пустые. В этой книге
формулы стоят ровно в трёх местах — себестоимость и FC % на листе «Меню», колонка «Сумма»
у поставщиков, «итого часов» и «рублей» в графике (406 ячеек). Все три ТЗ 7 прямо
предписывает парсеру не импортировать, поэтому для эталона формата потери нет. Ни одна
ячейка, нужная парсерам, формулой не является — проверено сверкой с исходником.

Запуск::

    .venv/bin/python tests/tools/sanitize_reference.py \\
        tests/fixtures/reference/Invasion.original.xlsx \\
        tests/fixtures/reference/Invasion.xlsx

Исходник остаётся локально и в git не попадает (см. .gitignore).
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

#: Телефон в любом из написаний, встречающихся в книге: +7 (985) 127 - 54 - 73,
#: +7 ( 912 ) 874 - 12 - 24, 8-999-1234567. Совпадение обязано кончаться цифрой,
#: иначе жадный класс съедает пробел перед следующим словом и склеивает текст.
PHONE: Final = re.compile(r"(?:\+7|\b8)[\s\-()0-9]{7,}[0-9]")

#: Строка контакта поставщика: «<имя> <телефон> / время заказа до 16:00».
#: Заменяется целиком — имя контактного лица тоже персональные данные.
CONTACT_REPLACEMENT: Final = "Контакт +7 (000) 000 - 00 - 00 / время заказа до 16:00"

#: Заголовки денежных колонок листа «Меню». Сравнение по нормализованному тексту,
#: потому что в книге они записаны капслоком и с опечаткой «НАИМЕНОВНИЕ».
MENU_MONEY_HEADERS: Final = frozenset(
    {
        "цена за порцию",
        "цена со скидкой",
        "себестоимость за порцию",
        "себестоимость",
        "fc %",
        "fc%",
    }
)

#: Служебные денежные колонки листа «ГРАФИК» — зарплатный контур, который по ТЗ 4.2
#: в бот не переносится вовсе.
SCHEDULE_MONEY_HEADERS: Final = frozenset({"ставка", "рублей"})


def normalise(value: object) -> str:
    """Нормализованный заголовок: без регистра, без лишних пробелов и переносов."""
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def synthetic_price(row: int, column: int) -> float:
    """Детерминированная синтетическая цена.

    Тип и порядок величины сохраняются — парсеру важно, что в ячейке float, а не строка,
    и что значение похоже на цену. Конкретная сумма значения не имеет.
    """
    return round(100 + (row * 37 + column * 13) % 4900 + 0.5, 2)


def sanitize_supplier_orders(ws: Worksheet) -> int:
    """Лист «ЗаказАлкогольSoft»: строки контактов и колонка «Цена»."""
    changed = 0
    price_column: int | None = None

    for row in range(1, ws.max_row + 1):
        first = ws.cell(row=row, column=1).value

        # Строка контакта: единственная ячейка с телефоном. Заменяется целиком,
        # но формат сохраняется — парсер должен вытащить из неё телефон и время приёма.
        if isinstance(first, str) and PHONE.search(first):
            ws.cell(row=row, column=1).value = CONTACT_REPLACEMENT
            changed += 1
            continue

        # Строка заголовков блока поставщика переопределяет положение колонки «Цена»:
        # блоки идут подряд и в принципе могут разъехаться по колонкам.
        if normalise(first) == "название":
            price_column = None
            for column in range(1, ws.max_column + 1):
                if normalise(ws.cell(row=row, column=column).value) == "цена":
                    price_column = column
                    break
            continue

        if price_column is not None:
            cell = ws.cell(row=row, column=price_column)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.value = synthetic_price(row, price_column)
                changed += 1

    return changed


def sanitize_by_header(ws: Worksheet, headers: frozenset[str], *, clear: bool) -> int:
    """Найти колонки по заголовку и обезличить их значения.

    ``clear=True`` — стереть (зарплатный контур, он не нужен вовсе).
    ``clear=False`` — заменить на синтетическое число (цены, тип ячейки важен парсеру).
    """
    targets: set[int] = set()
    for row in range(1, min(ws.max_row, 10) + 1):
        for column in range(1, ws.max_column + 1):
            if normalise(ws.cell(row=row, column=column).value) in headers:
                targets.add(column)

    changed = 0
    for column in sorted(targets):
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=column)
            if cell.value is None or normalise(cell.value) in headers:
                continue
            if not isinstance(cell.value, (int, float)) or isinstance(cell.value, bool):
                continue
            cell.value = None if clear else synthetic_price(row, column)
            changed += 1

    return changed


def sanitize(source: Path, target: Path) -> None:
    workbook = load_workbook(source)

    report: list[tuple[str, int]] = []

    orders = "ЗаказАлкогольSoft"
    if orders in workbook.sheetnames:
        report.append((orders, sanitize_supplier_orders(workbook[orders])))

    menu = "Меню"
    if menu in workbook.sheetnames:
        report.append((menu, sanitize_by_header(workbook[menu], MENU_MONEY_HEADERS, clear=False)))

    schedule = "ГРАФИК"
    if schedule in workbook.sheetnames:
        changed = sanitize_by_header(workbook[schedule], SCHEDULE_MONEY_HEADERS, clear=True)
        report.append((schedule, changed))

    # Последняя сеть: телефон, забытый в любой другой ячейке любого листа.
    leftovers = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and PHONE.search(cell.value):
                    cell.value = PHONE.sub("+7 (000) 000 - 00 - 00", cell.value)
                    leftovers += 1
    report.append(("телефоны на прочих листах", leftovers))

    workbook.save(target)

    for sheet, count in report:
        print(f"  {sheet:32} обезличено ячеек: {count}")


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    source, target = Path(sys.argv[1]), Path(sys.argv[2])
    if not source.exists():
        print(f"Нет исходника: {source}")
        return 1
    print(f"Обезличивание {source.name} -> {target.name}")
    sanitize(source, target)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
